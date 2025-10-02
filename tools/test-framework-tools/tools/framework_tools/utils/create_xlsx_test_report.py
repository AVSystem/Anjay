# -*- coding: utf-8 -*-
#
# Copyright 2017-2026 AVSystem <avsystem@avsystem.com>
# AVSystem Anjay LwM2M SDK
# All rights reserved.
#
# Licensed under AVSystem Anjay LwM2M Client SDK - Non-Commercial License.
# See the attached LICENSE file for details.

import argparse
import enum
import os
import re
import sys
import collections
from operator import itemgetter

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side


ColumnValues = collections.namedtuple("ColumnValues", ["int", "name"])

class ColumnIndex(enum.Enum):
    TYPE        = ColumnValues(1, "TYPE")
    SUITE       = ColumnValues(2, "SUITE")
    NAME        = ColumnValues(3, "NAME")
    STATUS      = ColumnValues(4, "STATUS")
    TIME        = ColumnValues(5, "TIME [s]")
    FAIL_REASON = ColumnValues(6, "FAIL REASON")

    def to_letter(self):
        return chr(ord('A') - 1 + self.value.int)

# Assert that the FAIL_REASON has the largest value
assert all(ColumnIndex.FAIL_REASON.value.int >= member.value.int for member in ColumnIndex), \
    "FAIL_REASON should be the largest value"

# Assert that integer values are in ascending order (1, 2, 3 ...)
for i, elem in enumerate(ColumnIndex):
    assert(elem.value.int == i + 1), \
        "ColumnIndex integer values should be arranged in ascending order"


class TestResultParser():
    def __init__(self, log_file_path) -> None:
        self.log_file_path = log_file_path

    def get_test_results(self):
        test_type, test_suite = self._get_test_type_and_suite()
        file_content = self._read_log_file_content()
        if file_content is None or len(file_content) == 0:
            return None
        test_name, test_status, test_time = self._get_result_info(file_content)
        fail_reason = self._get_fail_reason(test_status, file_content)

        ret_result = [None] * len(ColumnIndex)
        ret_result[ColumnIndex.TYPE.value.int - 1] = test_type
        ret_result[ColumnIndex.SUITE.value.int - 1] = test_suite
        ret_result[ColumnIndex.NAME.value.int - 1] = test_name
        ret_result[ColumnIndex.STATUS.value.int - 1] = test_status
        ret_result[ColumnIndex.TIME.value.int - 1] = test_time
        ret_result[ColumnIndex.FAIL_REASON.value.int - 1] = fail_reason
        return ret_result

    def _status_is_fail_or_error(self, status):
        if status == "FAIL" or status == "ERROR":
            return True
        return False

    def _read_log_file_content(self):
        if not os.path.exists(self.log_file_path):
            return None
        with open(self.log_file_path, 'r') as file:
            lines = file.readlines()

        return lines

    def _get_result_info(self, file_content):
        # an example success log file may look like this:
            #  SmsDtls  . . . . . . . . . . . . . . . . . . . . . . . . . . . OK (34.34s)
        # an example FAIL/ERROR log file may look like this:
            #  NonconfirmableExecuteTest  . . . . . . . . . . . . . . . . . . FAIL
            # Test NonconfirmableExecuteTest failed!
            # (assertion description...)
            #
            # Stack trace:
            # (python stack trace...)
        result_line = file_content[0].replace('. ', '').replace("  ", " ").strip().split(" ")
        test_name = result_line[0]
        test_status = result_line[1]
        try:
            test_time = result_line[2][1:-2]
        except:
            test_time = None

        return test_name, test_status, test_time

    def _has_whitespaces_only(self, line):
        if re.match(r'^\s*$', line):
            return True
        return False

    def _get_fail_reason(self, status, file_content):
        if not self._status_is_fail_or_error(status):
            return ""
        ret = []
        for line in file_content:
            if line == "Stack trace:\n":
                break
            if self._has_whitespaces_only(line):
                continue
            ret.append(line)

        return ''.join(ret[1:]).strip()

    def _get_test_type_and_suite(self):
        directory_name = os.path.basename(os.path.dirname(self.log_file_path)).split('.')
        test_type = directory_name[0]
        test_suite = '.'.join(directory_name[1:])

        return test_type, test_suite


class XlsxFileCreator():
    def __init__(self) -> None:
        self.wb = Workbook()
        self.ws = self.wb.active
        self.number_of_rows = 0

    def append_test_results(self, test_results):
        sorted_results = self._sort_test_results(test_results)
        for result in sorted_results:
            if result is not None:
                self.ws.append(result)
                self.number_of_rows += 1

    def perform_formating(self):
        self._apply_vertical_center_alignment()
        self._format_first_row()
        self._adjust_rows_height()
        self._adjust_columns_width()
        self._add_status_colors()
        self._format_time_as_numbers()
        self._add_borders()
        self._format_as_table()

    def add_header(self):
        header = []
        for elem in ColumnIndex:
            header.append(elem.value.name)
        self.ws.append(header)
        self.number_of_rows += 1

    def _sort_test_results(self, unsorted_results):
        sorted_results = sorted(unsorted_results, key=itemgetter(ColumnIndex.TYPE.value.int - 1,
                                                                 ColumnIndex.SUITE.value.int - 1,
                                                                 ColumnIndex.NAME.value.int - 1))
        return sorted_results

    def _adjust_rows_height(self):
        for row in self.ws.iter_rows():
            max_height = 0
            for cell in row:
                try:
                    # NOTE: partially based on:
                    # https://stackoverflow.com/questions/39529662/python-automatically-adjust-width-of-an-excel-files-columns
                    cell_height = len(str(cell.value).split('\n')) * 13 * cell.font.size/10
                    if cell_height > max_height:
                        max_height = cell_height
                except:
                    pass
            self.ws.row_dimensions[row[0].row].height = max_height

    def _adjust_columns_width(self):
        for column in self.ws.iter_cols():
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    # NOTE: partially based on:
                    # https://stackoverflow.com/questions/39529662/python-automatically-adjust-width-of-an-excel-files-columns
                    row_width = len(cell.value) * cell.font.size/10
                    if row_width > max_length:
                        max_length = row_width
                        if cell.font.bold == True:
                            max_length *= 1.1
                except:
                    pass
            adjusted_width = (max_length + 3)
            self.ws.column_dimensions[column_letter].width = adjusted_width

    def _apply_vertical_center_alignment(self):
        for row in self.ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical='center')
        for col in self.ws.iter_cols(min_col=1, max_col=len(ColumnIndex) - 1):
            for cell in col:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    def _format_first_row(self):
        fill = PatternFill(start_color="729FCF", end_color="729FCF", fill_type="solid")
        for cell in self.ws[1]:
            cell.font = Font(size=15, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = fill

    def _add_status_colors(self):
        ok_fill   = PatternFill(start_color="A1FF41", end_color="A1FF41", fill_type="solid")
        fail_fill = PatternFill(start_color="E11941", end_color="E11941", fill_type="solid")
        skip_fill = PatternFill(start_color="FDF362", end_color="FDF362", fill_type="solid")
        for col in self.ws.iter_cols(min_col=ColumnIndex.STATUS.value.int, max_col=ColumnIndex.STATUS.value.int):
            for cell in col:
                if cell.value == "OK":
                    cell.fill = ok_fill
                elif cell.value == "FAIL" or cell.value == "ERROR":
                    cell.fill = fail_fill
                elif cell.value == "SKIP":
                    cell.fill = skip_fill

    def _add_borders(self):
        border = Border(
            left   = Side(border_style="thin", color="000000"),
            right  = Side(border_style="thin", color="000000"),
            top    = Side(border_style="thin", color="000000"),
            bottom = Side(border_style="thin", color="000000"),
        )
        for row in self.ws.iter_rows():
            for cell in row:
                cell.border = border

    def _format_time_as_numbers(self):
        for col in self.ws.iter_cols(min_col=ColumnIndex.TIME.value.int, max_col=ColumnIndex.TIME.value.int):
            for cell in col:
                if cell.value != ColumnIndex.TIME.value.name and cell.value != None:
                    cell.number_format = "0.00"
                    cell.value = float(cell.value)

    def _format_as_table(self):
        ref = "A1:" + list(ColumnIndex)[-1].to_letter() + str(self.number_of_rows)
        table = Table(displayName="Table", ref=ref)

        style = TableStyleInfo(
            name="TableStyleMedium9", showFirstColumn=False,
            showLastColumn=False, showRowStripes=False, showColumnStripes=False
        )
        table.tableStyleInfo = style

        self.ws.add_table(table)


def enumerate_log_files(path):
    for root, _, files in os.walk(path, topdown=False):
        for file in files:
            if re.match(r'.*\.log$', file):
                yield os.path.join(root, file)


def _main():
    parser = argparse.ArgumentParser(description='Create .xlsx file out of output .log files',
                                     formatter_class=argparse.ArgumentDefaultsHelpFormatter)
    parser.add_argument('-d', '--log-dir',
                        help='directory that contains output .log files',
                        required=True)
    args = parser.parse_args()

    test_results = []
    for log_file_path in enumerate_log_files(args.log_dir):
        result_parser = TestResultParser(log_file_path)
        test_results.append(result_parser.get_test_results())

    xlsx_file = XlsxFileCreator()
    xlsx_file.add_header()
    xlsx_file.append_test_results(test_results)

    xlsx_file.perform_formating()

    output_path = os.path.join(args.log_dir, "../..", "integration_test_results.xlsx")
    xlsx_file.wb.save(output_path)

    print(f"{os.path.basename(__file__)}: saved test reports to {os.path.abspath(output_path)}")

if __name__ == '__main__':
    sys.exit(_main())
